from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, session
from app.auth.routes import login_required
import pyodbc
from datetime import datetime, date
from calendar import monthrange
from config import config

financial = Blueprint('financial', __name__)

# 매입/매출 집계
@financial.route('/financial_summary')
@login_required
def financial_summary():
    # 기본값 설정: 당월 1일부터 마지막일까지
    today = date.today()
    first_day = date(today.year, today.month, 1)
    last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
    
    start_date = request.args.get('start_date', first_day.strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', last_day.strftime('%Y-%m-%d'))
    
    try:
        conn = pyodbc.connect(config['development'].PYODBC_CONN_STR)
        cursor = conn.cursor()
        
        # Get purchase summary
        purchase_query = """
            SELECT 
                SUM(p.total_amount) as total_purchase,
                SUM(COALESCE(pp.amount, 0)) as total_purchase_payment
            FROM purchase_master p
            LEFT JOIN purchase_payment pp ON p.purchase_no = pp.purchase_no
            WHERE 1=1
        """
        purchase_params = []
        
        if start_date:
            purchase_query += " AND p.purchase_date >= ?"
            purchase_params.append(start_date)
        if end_date:
            purchase_query += " AND p.purchase_date <= ?"
            purchase_params.append(end_date)
        
        cursor.execute(purchase_query, purchase_params)
        purchase_summary = cursor.fetchone()
        
        # Get sales summary
        sales_query = """
            SELECT 
                SUM(s.total_amount) as total_sales,
                SUM(COALESCE(sp.amount, 0)) as total_sales_payment
            FROM sales_master s
            LEFT JOIN sales_payment sp ON s.sales_no = sp.sales_no
            WHERE 1=1
        """
        sales_params = []
        
        if start_date:
            sales_query += " AND s.sales_date >= ?"
            sales_params.append(start_date)
        if end_date:
            sales_query += " AND s.sales_date <= ?"
            sales_params.append(end_date)
        
        cursor.execute(sales_query, sales_params)
        sales_summary = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return render_template('financial/summary.html',
                             purchase_summary=purchase_summary,
                             sales_summary=sales_summary,
                             default_start_date=first_day.strftime('%Y-%m-%d'),
                             default_end_date=last_day.strftime('%Y-%m-%d'))
    except Exception as e:
        flash(f'Database error: {str(e)}', 'error')
        return render_template('financial/summary.html',
                             purchase_summary=None,
                             sales_summary=None,
                             default_start_date=first_day.strftime('%Y-%m-%d'),
                             default_end_date=last_day.strftime('%Y-%m-%d'))

# 결산마감
@financial.route('/financial_close', methods=['GET', 'POST'])
@login_required
def financial_close():
    if request.method == 'POST':
        close_date = request.form['close_date']
        
        try:
            conn = pyodbc.connect(config['development'].PYODBC_CONN_STR)
            cursor = conn.cursor()
            
            # Check if already closed
            cursor.execute("SELECT * FROM financial_close WHERE close_date = ?", (close_date,))
            if cursor.fetchone():
                flash('This period is already closed', 'warning')
                return redirect(url_for('financial.financial_close'))
            
            # Get the latest closed date
            cursor.execute("SELECT MAX(close_date) as latest_close_date FROM financial_close")
            latest_close = cursor.fetchone()
            latest_close_date = latest_close.latest_close_date if latest_close and latest_close.latest_close_date else None
            
            # Check if close_date is after the latest closed date
            if latest_close_date and close_date <= latest_close_date.strftime('%Y-%m-%d'):
                flash(f'Cannot close a period on or before the latest closed date ({latest_close_date.strftime("%Y-%m-%d")})', 'warning')
                return redirect(url_for('financial.financial_close'))
            
            # Get purchase summary for the period
            if latest_close_date:
                # 이전 마감일 이후부터 현재 마감일까지의 데이터만 집계
                cursor.execute("""
                    SELECT 
                        SUM(p.total_amount) as total_purchase,
                        SUM(COALESCE(pp.amount, 0)) as total_purchase_payment
                    FROM purchase_master p
                    LEFT JOIN purchase_payment pp ON p.purchase_no = pp.purchase_no
                    WHERE p.purchase_date > ? AND p.purchase_date <= ?
                """, (latest_close_date.strftime('%Y-%m-%d'), close_date))
            else:
                # 첫 번째 마감인 경우 마감일까지의 모든 데이터 집계
                cursor.execute("""
                    SELECT 
                        SUM(p.total_amount) as total_purchase,
                        SUM(COALESCE(pp.amount, 0)) as total_purchase_payment
                    FROM purchase_master p
                    LEFT JOIN purchase_payment pp ON p.purchase_no = pp.purchase_no
                    WHERE p.purchase_date <= ?
                """, (close_date,))
            purchase_summary = cursor.fetchone()
            
            # Get sales summary for the period
            if latest_close_date:
                # 이전 마감일 이후부터 현재 마감일까지의 데이터만 집계
                cursor.execute("""
                    SELECT 
                        SUM(s.total_amount) as total_sales,
                        SUM(COALESCE(sp.amount, 0)) as total_sales_payment
                    FROM sales_master s
                    LEFT JOIN sales_payment sp ON s.sales_no = sp.sales_no
                    WHERE s.sales_date > ? AND s.sales_date <= ?
                """, (latest_close_date.strftime('%Y-%m-%d'), close_date))
            else:
                # 첫 번째 마감인 경우 마감일까지의 모든 데이터 집계
                cursor.execute("""
                    SELECT 
                        SUM(s.total_amount) as total_sales,
                        SUM(COALESCE(sp.amount, 0)) as total_sales_payment
                    FROM sales_master s
                    LEFT JOIN sales_payment sp ON s.sales_no = sp.sales_no
                    WHERE s.sales_date <= ?
                """, (close_date,))
            sales_summary = cursor.fetchone()
            
            # Record financial close
            cursor.execute("""
                INSERT INTO financial_close (
                    close_date, total_purchase, total_purchase_payment,
                    total_sales, total_sales_payment, closed_by, closed_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                close_date,
                purchase_summary.total_purchase or 0,
                purchase_summary.total_purchase_payment or 0,
                sales_summary.total_sales or 0,
                sales_summary.total_sales_payment or 0,
                session.get('username', 'Unknown'),
                datetime.now()
            ))
            
            conn.commit()
            cursor.close()
            conn.close()
            flash('Financial period closed successfully', 'success')
            return redirect(url_for('financial.financial_close'))
        except Exception as e:
            flash(f'Database error: {str(e)}', 'error')
            return redirect(url_for('financial.financial_close'))
    
    try:
        conn = pyodbc.connect(config['development'].PYODBC_CONN_STR)
        cursor = conn.cursor()
        
        # 페이지네이션 설정
        page = request.args.get('page', 1, type=int)
        per_page = 10
        offset = (page - 1) * per_page
        
        # 전체 마감 기간 수 조회
        cursor.execute("SELECT COUNT(*) as total FROM financial_close")
        total_count = cursor.fetchone().total
        
        # Get closed periods with pagination
        cursor.execute("""
            SELECT close_date, total_purchase, total_purchase_payment,
                   total_sales, total_sales_payment, closed_by, closed_at
            FROM financial_close
            ORDER BY close_date DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, (offset, per_page))
        closed_periods = cursor.fetchall()
        
        # 전체 합계 계산
        cursor.execute("""
            SELECT 
                SUM(total_purchase) as total_purchase_sum,
                SUM(total_purchase_payment) as total_purchase_payment_sum,
                SUM(total_sales) as total_sales_sum,
                SUM(total_sales_payment) as total_sales_payment_sum
            FROM financial_close
        """)
        totals = cursor.fetchone()
        
        # Get the latest closed date to set minimum close date
        cursor.execute("SELECT MAX(close_date) as latest_close_date FROM financial_close")
        latest_close = cursor.fetchone()
        latest_close_date = latest_close.latest_close_date if latest_close and latest_close.latest_close_date else None
        
        cursor.close()
        conn.close()
        
        # 페이지네이션 정보 계산
        total_pages = (total_count + per_page - 1) // per_page
        
        # 기본값을 최신 마감일 다음날로 설정 (마감일이 없으면 당일)
        if latest_close_date:
            from datetime import timedelta
            next_day = latest_close_date + timedelta(days=1)
            default_close_date = next_day.strftime('%Y-%m-%d')
        else:
            default_close_date = date.today().strftime('%Y-%m-%d')
        
        return render_template('financial/close.html', 
                             closed_periods=closed_periods,
                             totals=totals,
                             pagination={
                                 'page': page,
                                 'per_page': per_page,
                                 'total_count': total_count,
                                 'total_pages': total_pages
                             },
                             default_close_date=default_close_date,
                             latest_close_date=latest_close_date.strftime('%Y-%m-%d') if latest_close_date else None)
    except Exception as e:
        flash(f'Database error: {str(e)}', 'error')
        return render_template('financial/close.html', 
                             closed_periods=[],
                             default_close_date=date.today().strftime('%Y-%m-%d'))

# 마감 취소
@financial.route('/financial_close_cancel/<close_date>', methods=['POST'])
@login_required
def financial_close_cancel(close_date):
    try:
        conn = pyodbc.connect(config['development'].PYODBC_CONN_STR)
        cursor = conn.cursor()
        
        # Check if period exists
        cursor.execute("SELECT * FROM financial_close WHERE close_date = ?", (close_date,))
        if not cursor.fetchone():
            flash('Period not found', 'error')
            return redirect(url_for('financial.financial_close'))
        
        # Delete the closed period
        cursor.execute("DELETE FROM financial_close WHERE close_date = ?", (close_date,))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        flash(f'Financial period {close_date} has been cancelled successfully', 'success')
        return redirect(url_for('financial.financial_close'))
    except Exception as e:
        flash(f'Database error: {str(e)}', 'error')
        return redirect(url_for('financial.financial_close')) 

# 손익계산서 P&L Report
@financial.route('/pnl_report')
@login_required
def pnl_report():
    from collections import OrderedDict
    year = request.args.get('year', datetime.now().year, type=int)

    def _norm(d):
        return {m: float(d.get(m, 0)) for m in range(1, 13)}

    def _code_agg(rows):
        d = OrderedDict()
        for r in rows:
            code, mo, amt = r[0], r[1], float(r[2] or 0)
            if code not in d:
                d[code] = {m: 0.0 for m in range(1, 13)}
            d[code][mo] = amt
        return d

    def _stats(d):
        total = sum(d.values())
        cnt   = sum(1 for v in d.values() if v != 0)
        avg   = total / cnt if cnt else 0.0
        return {'total': total, 'cnt': cnt, 'avg': avg}

    empty = {m: 0.0 for m in range(1, 13)}
    dummy = {'total': 0.0, 'cnt': 0, 'avg': 0.0}

    try:
        conn   = pyodbc.connect(config['development'].PYODBC_CONN_STR)
        cursor = conn.cursor()

        # 코드 이름 조회 (세션 언어에 따라)
        lang_raw = (session.get('lang') or session.get('language', 'ko')).lower()
        lang_map = {
            'ko': 'name_ko', 'korean': 'name_ko',
            'en': 'name_en', 'english': 'name_en',
            'th': 'name_th', 'thai': 'name_th',
        }
        name_col = lang_map.get(lang_raw, 'name_ko')
        cursor.execute(f"SELECT account_code, {name_col} FROM code_Account")
        code_names = {str(r[0]): r[1] for r in cursor.fetchall()}

        # 1. Sales
        cursor.execute("SELECT MONTH(sales_date), SUM(total_amount) FROM sales_master WHERE YEAR(sales_date)=? GROUP BY MONTH(sales_date)", (year,))
        sales = _norm({r[0]: r[1] for r in cursor.fetchall()})

        # 2. Purchase
        cursor.execute("SELECT MONTH(purchase_date), SUM(total_amount) FROM purchase_master WHERE YEAR(purchase_date)=? GROUP BY MONTH(purchase_date)", (year,))
        purchase = _norm({r[0]: r[1] for r in cursor.fetchall()})

        # 3. C = A - B
        c_monthly = {m: sales[m] - purchase[m] for m in range(1, 13)}

        # 4. Labor Cost
        labor = OrderedDict()
        
        # 4-1. Regular Salary (from salary_log table)
        cursor.execute("""
            SELECT Month, SUM(PaidAmount) 
            FROM salary_log 
            WHERE Year = ? AND CompanyName = 'RECYCLE'
            GROUP BY Month
        """, (year,))
        salary_data = _norm({r[0]: r[1] for r in cursor.fetchall()})
        labor['salary'] = salary_data
        
        # 4-2. Parttime (from Account_Actual, COST_CENTER='RECYCLE', Actual_Code='5150')
        cursor.execute("""
            SELECT MONTH(BILLING_DATE), SUM(BEFORE_VAT_AMT)
            FROM Account_Actual
            WHERE YEAR(BILLING_DATE) = ?
              AND COST_CENTER = 'RECYCLE'
              AND Actual_Code = '5150'
            GROUP BY MONTH(BILLING_DATE)
        """, (year,))
        parttime_data = _norm({r[0]: r[1] for r in cursor.fetchall()})
        labor['parttime'] = parttime_data
        
        # Labor Sum
        labor_sum = {m: sum(d[m] for d in labor.values()) for m in range(1, 13)}

        # 5. Expense (COST_CENTER='RECYCLE', Actual_Code NOT IN ('5150'))
        cursor.execute("""
            SELECT Actual_Code, MONTH(BILLING_DATE), SUM(BEFORE_VAT_AMT)
            FROM Account_Actual
            WHERE YEAR(BILLING_DATE) = ?
              AND COST_CENTER = 'RECYCLE'
              AND Actual_Code != '5150'
            GROUP BY Actual_Code, MONTH(BILLING_DATE)
            ORDER BY Actual_Code
        """, (year,))
        expense     = _code_agg(cursor.fetchall())
        expense_sum = {m: sum(d[m] for d in expense.values()) for m in range(1, 13)}

        # 6. C - Labor Cost Sum
        c_minus_labor = {m: c_monthly[m] - labor_sum[m] for m in range(1, 13)}

        # 7. Net1 = (C - Labor Cost Sum) - Expense Sum
        net1  = {m: c_minus_labor[m] - expense_sum[m] for m in range(1, 13)}
        final = net1

        cursor.close()
        conn.close()

        return render_template('financial/pnl.html',
            year=year,
            code_names=code_names,
            sales=sales,            sales_stats=_stats(sales),
            purchase=purchase,      purchase_stats=_stats(purchase),
            c_monthly=c_monthly,    c_stats=_stats(c_monthly),
            labor=labor,            labor_sum=labor_sum,
            labor_stats={code: _stats(d) for code, d in labor.items()},
            labor_sum_stats=_stats(labor_sum),
            c_minus_labor=c_minus_labor, c_minus_labor_stats=_stats(c_minus_labor),
            expense=expense,        expense_sum=expense_sum,
            expense_stats={code: _stats(d) for code, d in expense.items()},
            expense_sum_stats=_stats(expense_sum),
            net1=net1,              net1_stats=_stats(net1),
            final=final,            final_stats=_stats(final),
        )
    except Exception as e:
        flash(f'Database error: {str(e)}', 'error')
        from collections import OrderedDict
        return render_template('financial/pnl.html',
            year=year,
            code_names={},
            sales=empty, sales_stats=dummy, purchase=empty, purchase_stats=dummy,
            c_monthly=empty, c_stats=dummy,
            labor=OrderedDict(), labor_sum=empty, labor_stats={}, labor_sum_stats=dummy,
            c_minus_labor=empty, c_minus_labor_stats=dummy,
            expense=OrderedDict(), expense_sum=empty, expense_stats={}, expense_sum_stats=dummy,
            net1=empty, net1_stats=dummy,
            final=empty, final_stats=dummy,
        )

@financial.route('/pnl_report/excel')
@login_required
def pnl_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import make_response
    from collections import OrderedDict
    import io

    year = request.args.get('year', datetime.now().year, type=int)

    def _norm(d):
        return {m: float(d.get(m, 0)) for m in range(1, 13)}

    def _code_agg(rows):
        d = OrderedDict()
        for r in rows:
            code, mo, amt = r[0], r[1], float(r[2] or 0)
            if code not in d:
                d[code] = {m: 0.0 for m in range(1, 13)}
            d[code][mo] = amt
        return d

    def _stats(d):
        total = sum(d.values())
        cnt   = sum(1 for v in d.values() if v != 0)
        avg   = total / cnt if cnt else 0.0
        return {'total': total, 'cnt': cnt, 'avg': avg}

    conn   = pyodbc.connect(config['development'].PYODBC_CONN_STR)
    cursor = conn.cursor()

    lang     = session.get('lang') or session.get('language', 'ko')
    lang     = lang.lower()
    name_col = {'ko':'name_ko','korean':'name_ko','en':'name_en','english':'name_en','th':'name_th','thai':'name_th'}.get(lang, 'name_ko')
    cursor.execute(f"SELECT account_code, {name_col} FROM code_Account")
    code_names = {str(r[0]): r[1] for r in cursor.fetchall()}

    cursor.execute("SELECT MONTH(sales_date), SUM(total_amount) FROM sales_master WHERE YEAR(sales_date)=? GROUP BY MONTH(sales_date)", (year,))
    sales    = _norm({r[0]: r[1] for r in cursor.fetchall()})

    cursor.execute("SELECT MONTH(purchase_date), SUM(total_amount) FROM purchase_master WHERE YEAR(purchase_date)=? GROUP BY MONTH(purchase_date)", (year,))
    purchase = _norm({r[0]: r[1] for r in cursor.fetchall()})

    c_monthly = {m: sales[m] - purchase[m] for m in range(1, 13)}

    # Labor Cost
    labor = OrderedDict()
    
    # Regular Salary
    cursor.execute("""
        SELECT Month, SUM(PaidAmount) 
        FROM salary_log 
        WHERE Year = ? AND CompanyName = 'RECYCLE'
        GROUP BY Month
    """, (year,))
    labor['salary'] = _norm({r[0]: r[1] for r in cursor.fetchall()})
    
    # Parttime
    cursor.execute("""
        SELECT MONTH(BILLING_DATE), SUM(BEFORE_VAT_AMT)
        FROM Account_Actual
        WHERE YEAR(BILLING_DATE) = ?
          AND COST_CENTER = 'RECYCLE'
          AND Actual_Code = '5150'
        GROUP BY MONTH(BILLING_DATE)
    """, (year,))
    labor['parttime'] = _norm({r[0]: r[1] for r in cursor.fetchall()})
    
    labor_sum = {m: sum(d[m] for d in labor.values()) for m in range(1, 13)}

    # Expense
    cursor.execute("""
        SELECT Actual_Code, MONTH(BILLING_DATE), SUM(BEFORE_VAT_AMT)
        FROM Account_Actual
        WHERE YEAR(BILLING_DATE) = ?
          AND COST_CENTER = 'RECYCLE'
          AND Actual_Code != '5150'
        GROUP BY Actual_Code, MONTH(BILLING_DATE)
        ORDER BY Actual_Code
    """, (year,))
    expense     = _code_agg(cursor.fetchall())
    expense_sum = {m: sum(d[m] for d in expense.values()) for m in range(1, 13)}
    net1        = {m: c_monthly[m] - expense_sum[m] for m in range(1, 13)}
    final       = net1

    cursor.close()
    conn.close()

    # ── 엑셀 생성 (기존 로직 유지) ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"P&L {year}"

    # 스타일 정의
    hdr_fill   = PatternFill("solid", fgColor="2D3748")
    hdr_font   = Font(color="FFFFFF", bold=True, size=10)
    year_fill  = PatternFill("solid", fgColor="4A5568")
    labor_fill = PatternFill("solid", fgColor="FFFDE7")
    exp_fill   = PatternFill("solid", fgColor="EBF8FF")
    net_fill   = PatternFill("solid", fgColor="F0FFF4")
    final_fill = PatternFill("solid", fgColor="2D3748")
    final_font = Font(color="FFD700", bold=True, size=10)
    sum_fill   = PatternFill("solid", fgColor="EDF2F7")
    center     = Alignment(horizontal="center", vertical="center")
    right      = Alignment(horizontal="right",  vertical="center")
    left       = Alignment(horizontal="left",   vertical="center")
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fmt(v):
        # 숫자를 그대로 반환 (엑셀 number_format으로 표시)
        if v is None or v == 0: 
            return None
        return float(v)

    def cell(row, col, value, fill=None, font=None, align=None, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        if fill:  c.fill  = fill
        if font:  c.font  = font
        elif bold: c.font = Font(bold=True, size=10)
        else:      c.font = Font(size=10)
        c.alignment = align or center
        c.border    = border
        return c

    # 행 1: 타이틀 헤더 (Item | 2026 × 14)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
    cell(1, 1, "Item", hdr_fill, hdr_font, center)

    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=16)
    cell(1, 3, str(year), year_fill, hdr_font, center)

    # 행 2: 월 헤더
    for i, m in enumerate(range(1, 13), start=3):
        cell(2, i, m, hdr_fill, hdr_font, center)
    cell(2, 15, "Sum", hdr_fill, hdr_font, center)
    cell(2, 16, "Avg", hdr_fill, hdr_font, center)

    r = 3

    def write_row(label, data_dict, stats, row, fill=None, bold=False, neg_color=False, colspan_item=True):
        col_start = 1 if colspan_item else 2
        if colspan_item:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=col_start, value=label)
        c.fill = fill or PatternFill()
        c.font = Font(bold=bold, size=10)
        c.alignment = left
        c.border = border
        if colspan_item:
            ws.cell(row=row, column=2).border = border

        for i, m in enumerate(range(1, 13), start=3):
            v = data_dict.get(m, 0)
            val = fmt(v)
            wc = ws.cell(row=row, column=i, value=val)
            wc.fill = fill or PatternFill()
            wc.font = Font(size=10, color="C53030" if (neg_color and v < 0) else "000000")
            wc.alignment = right
            wc.border = border
            if val is not None:
                wc.number_format = '#,##0.00;(#,##0.00);"-"'

        # Sum
        sum_val = fmt(stats['total'])
        sc = ws.cell(row=row, column=15, value=sum_val)
        sc.fill = sum_fill
        sc.font = Font(bold=True, size=10, color="C53030" if (neg_color and stats['total'] < 0) else "000000")
        sc.alignment = right
        sc.border = border
        if sum_val is not None:
            sc.number_format = '#,##0.00;(#,##0.00);"-"'

        # Avg
        avg_val = fmt(stats['avg'])
        ac = ws.cell(row=row, column=16, value=avg_val)
        ac.fill = PatternFill("solid", fgColor="F7FAFC")
        ac.font = Font(size=10)
        ac.alignment = right
        ac.border = border
        if avg_val is not None:
            ac.number_format = '#,##0.00;(#,##0.00);"-"'

    # Sales
    write_row("Sales (A)", sales, _stats(sales), r)
    r += 1
    # Purchasing
    write_row("Purchasing (B)", purchase, _stats(purchase), r)
    r += 1
    # C = A-B
    write_row("C = A-B", c_monthly, _stats(c_monthly), r, neg_color=True)
    r += 1

    # Labor
    labor_codes = list(labor.keys())
    if labor_codes:
        section_start = r
        for i, code in enumerate(labor_codes):
            st    = _stats(labor[code])
            label = f"{code} - {code_names.get(code,'')}" if code_names.get(code) else code
            if i == 0:
                label += " (D)"
            elif i == 1:
                label += " (E)"
            # section 셀
            sc = ws.cell(row=r, column=1, value="Labor Cost" if i == 0 else None)
            sc.fill = PatternFill("solid", fgColor="F9A825")
            sc.font = Font(color="1a202c", bold=True, size=10)
            sc.alignment = center
            sc.border = border
            write_row(label, labor[code], st, r, fill=PatternFill("solid", fgColor="FFFDE7"), colspan_item=False)
            r += 1
        # Labor Sum 행 추가
        sc = ws.cell(row=r, column=1, value=None)
        sc.fill = PatternFill("solid", fgColor="F9A825")
        sc.border = border
        write_row("Labor Cost Sum (F = D + E)", labor_sum, _stats(labor_sum), r, fill=PatternFill("solid", fgColor="FFF9C4"), bold=True, colspan_item=False)
        r += 1
        ws.merge_cells(start_row=section_start, start_column=1, end_row=r-1, end_column=1)
    
    # C - Labor Cost Sum
    c_minus_labor = {m: c_monthly[m] - labor_sum[m] for m in range(1, 13)}
    write_row("(C - Labor Cost Sum)", c_minus_labor, _stats(c_minus_labor), r, neg_color=True)
    r += 1

    # Expense
    exp_codes = list(expense.keys())
    if exp_codes:
        section_start = r
        for i, code in enumerate(exp_codes):
            st    = _stats(expense[code])
            label = f"{code} - {code_names.get(code,'')}" if code_names.get(code) else code
            sc = ws.cell(row=r, column=1, value="Expense" if i == 0 else None)
            sc.fill = PatternFill("solid", fgColor="3182CE")
            sc.font = Font(color="FFFFFF", bold=True, size=10)
            sc.alignment = center
            sc.border = border
            write_row(label, expense[code], st, r, fill=PatternFill("solid", fgColor="FFFFFF"), colspan_item=False)
            r += 1
        # Expense Sum
        sc = ws.cell(row=r, column=1, value=None)
        sc.fill = PatternFill("solid", fgColor="3182CE")
        sc.border = border
        write_row("Expense Sum", expense_sum, _stats(expense_sum), r, fill=exp_fill, bold=True, colspan_item=False)
        r += 1
        ws.merge_cells(start_row=section_start, start_column=1, end_row=r-1, end_column=1)

    # Net1 = (C - Labor Cost Sum) - Expense Sum
    net1_calc = {m: c_minus_labor[m] - expense_sum[m] for m in range(1, 13)}
    write_row("((C - Labor Cost Sum) - Expense Sum)", net1_calc, _stats(net1_calc), r, fill=PatternFill("solid", fgColor="F0FFF4"), neg_color=True)
    r += 1

    # Final
    final = net1_calc

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=1, value=f"({year} Net)")
    c.fill = PatternFill("solid", fgColor="2D3748")
    c.font = Font(color="FFD700", bold=True, size=10)
    c.alignment = left
    c.border = border
    ws.cell(row=r, column=2).border = border

    for i, m in enumerate(range(1, 13), start=3):
        v = final.get(m, 0)
        val = fmt(v)
        wc = ws.cell(row=r, column=i, value=val)
        wc.fill = PatternFill("solid", fgColor="2D3748")
        wc.font = Font(color="FFD700", bold=True, size=10)
        wc.alignment = right
        wc.border = border
        if val is not None:
            wc.number_format = '#,##0.00;(#,##0.00);"-"'
    sum_val = fmt(_stats(final)['total'])
    sc = ws.cell(row=r, column=15, value=sum_val)
    sc.fill = PatternFill("solid", fgColor="2D3748")
    sc.font = Font(color="FFD700", bold=True, size=10)
    sc.alignment = right
    sc.border = border

    if sum_val is not None:
        sc.number_format = '#,##0.00;(#,##0.00);"-"'

    # 컬럼 너비
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(2)].width = 28
    for i in range(3, 16):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions[get_column_letter(16)].width = 14

    # 응답
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=PNL_{year}.xlsx'
    return resp