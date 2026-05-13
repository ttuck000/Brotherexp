from flask import Blueprint, render_template, jsonify, request, redirect, url_for, flash
from app.auth.routes import login_required
import pyodbc
from datetime import datetime
from flask_login import current_user
from app.inventory import bp as inventory

# Use the package-level `bp` imported above. Do NOT recreate a new Blueprint here
# (creating a second Blueprint with the same name caused endpoints to be attached
# to a different object and prevented `url_for('inventory.xxx')` from resolving.)

# 재고현황
@inventory.route('/inventory_status')
@login_required
def inventory_status():
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    
    # Get inventory status
    cursor.execute("""
        SELECT i.item_code, i.item_name, i.item_spec, i.item_unit,
               COALESCE(SUM(CASE WHEN t.transaction_type = 'IN' THEN t.quantity ELSE -t.quantity END), 0) as current_stock
        FROM item_code i
        LEFT JOIN transaction t ON i.item_code = t.item_code
        GROUP BY i.item_code, i.item_name, i.item_spec, i.item_unit
    """)
    inventory_status = cursor.fetchall()
    
    return render_template('inventory/inventory_status.html', inventory_status=inventory_status)

# 재고실사
@inventory.route('/inventory_check', methods=['GET', 'POST'])
@login_required
def inventory_check():
    if request.method == 'POST':
        item_code = request.form['item_code']
        actual_quantity = request.form['actual_quantity']
        check_date = datetime.now()
        
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        
        # Get current stock
        cursor.execute("""
            SELECT COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity ELSE -quantity END), 0)
            FROM transaction
            WHERE item_code = ?
        """, (item_code,))
        current_stock = cursor.fetchone()[0]
        
        # Record inventory check
        cursor.execute("""
            INSERT INTO inventory_check (item_code, check_date, system_quantity, actual_quantity)
            VALUES (?, ?, ?, ?)
        """, (item_code, check_date, current_stock, actual_quantity))
        
        # If there's a difference, create adjustment transaction
        if current_stock != actual_quantity:
            adjustment_quantity = actual_quantity - current_stock
            cursor.execute("""
                INSERT INTO transaction (item_code, transaction_date, transaction_type, quantity, remark)
                VALUES (?, ?, ?, ?, ?)
            """, (item_code, check_date, 'ADJ', adjustment_quantity, 'Inventory check adjustment'))
        
        conn.commit()
        flash('Inventory check completed successfully')
        return redirect(url_for('inventory.inventory_check'))
    
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM item_code")
    items = cursor.fetchall()
    
    return render_template('inventory/inventory_check.html', items=items)

# 수불현황
@inventory.route('/stock_status')
@login_required
def stock_status():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    item_code = request.args.get('item_code')
    transaction_date = request.args.get('transaction_date')
    
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    
    query = """
        SELECT t.transaction_date, t.item_code, i.item_name, i.item_spec, i.item_unit,
               t.transaction_type, t.quantity, t.remark
        FROM transaction t
        JOIN item_code i ON t.item_code = i.item_code
        WHERE 1=1
    """
    params = []
    
    if start_date:
        query += " AND t.transaction_date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND t.transaction_date <= ?"
        params.append(end_date)
    if item_code:
        query += " AND t.item_code = ?"
        params.append(item_code)
    if transaction_date:
        query += " AND t.transaction_date = ?"
        params.append(transaction_date)
    
    query += " ORDER BY t.transaction_date DESC"
    
    cursor.execute(query, params)
    transactions = cursor.fetchall()
    
    return render_template('inventory/stock_status.html', transactions=transactions)

@inventory.route('/inventory/inventory')
@login_required
def inventory_page():
    return render_template('inventory/inventory.html')

@inventory.route('/inventory/inventory_list')
@login_required
def inventory_list():
    return render_template('inventory/inventory.html')

@inventory.route('/inventory_transaction')
@login_required
def inventory_transaction():
    return render_template('inventory/inventory_transaction.html')

@inventory.route('/inventory/inventory_audit_list')
@login_required
def inventory_audit_list():
    return render_template('inventory/inventory_audit_list.html')

@inventory.route('/inventory_audit_register')
@login_required
def inventory_audit_register():
    return render_template('inventory/inventory_audit_register.html')

@inventory.route('/inventory/inventory_audit_edit')
@login_required
def inventory_audit_edit():
    return render_template('inventory/inventory_audit_edit.html')

@inventory.route('/inventory/inventory_audit_detail')
@login_required
def inventory_audit_detail_page():
    return render_template('inventory/inventory_audit_detail.html')

@inventory.route('/inventory/inventory_audit_detail_list')
@login_required
def inventory_audit_detail_list_page():
    return render_template('inventory/inventory_audit_detail_list.html')

@inventory.route('/api/warehouse/options')
@login_required
def warehouse_options():
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        cursor.execute('SELECT WH_CODE, WH_NAME FROM code_warehouse ORDER BY WH_NAME')
        rows = cursor.fetchall()
        return jsonify([{'code': r.WH_CODE, 'name': r.WH_NAME} for r in rows])
    except Exception as e:
        print('[창고 옵션 API 오류]', str(e))
        return jsonify([]), 500

@inventory.route('/api/inventory/item/options')
@login_required
def inventory_item_options():
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    cursor.execute('SELECT ItemCode, ItemName FROM ItemMaster ORDER BY ItemCode')
    rows = cursor.fetchall()
    return jsonify([{'item_code': r.ItemCode, 'item_name': r.ItemName} for r in rows])

@inventory.route('/api/inventory/item/detail/<item_code>')
@login_required
def inventory_item_detail(item_code):
    warehouse_code = request.args.get('warehouse_code')
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    cursor.execute('SELECT ItemCode, ItemName, Spec, unit FROM ItemMaster WHERE ItemCode = ?', (item_code,))
    row = cursor.fetchone()
    if not row:
        return jsonify({}), 404
    # 시스템재고 조회
    if warehouse_code:
        cursor.execute('SELECT CurrentStock FROM Inventory WHERE ItemCode = ? AND WarehouseCode = ?', (item_code, warehouse_code))
        inv_row = cursor.fetchone()
    else:
        cursor.execute('SELECT CurrentStock FROM Inventory WHERE ItemCode = ?', (item_code,))
        inv_row = cursor.fetchone()
    current_stock = float(inv_row.CurrentStock) if inv_row and inv_row.CurrentStock is not None else 0.0
    return jsonify({
        'item_code': row.ItemCode,
        'item_name': row.ItemName,
        'item_spec': row.Spec,
        'item_unit': row.unit,
        'system_stock': current_stock
    })

def safe_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0

def safe_str(val):
    return val if val not in (None, "") else None

@inventory.route('/api/inventory/audit/register', methods=['POST'])
@login_required
def inventory_audit_register_save():
    data = request.get_json()
    audit_date = data.get('auditDate')
    warehouse_code = data.get('warehouseCode')
    remarks = safe_str(data.get('remarks'))
    items = data.get('items', [])
    username = current_user.username if hasattr(current_user, 'username') else 'system'
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    try:
        print('[실사마스터 입력값]', audit_date, warehouse_code, remarks, username, now)
        # 마스터 저장 및 id 반환
        try:
            cursor.execute('''
                INSERT INTO inventory_audit_master (audit_date, warehouse_code, remarks, createuser, createdate, status)
                OUTPUT INSERTED.id
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (audit_date, warehouse_code, remarks, username, now, 'N'))
            row = cursor.fetchone()
            master_id = row[0] if row and row[0] is not None else None
            print('[실사마스터 INSERT 결과] master_id:', master_id)
        except Exception as e:
            print('[실사마스터 INSERT 오류]', str(e))
            conn.rollback()
            return jsonify({'success': False, 'message': '실사 마스터 저장에 실패했습니다.'})
        if not master_id:
            conn.rollback()
            return jsonify({'success': False, 'message': '실사 마스터 저장에 실패했습니다.'})
        # 상세 저장
        for idx, item in enumerate(items, 1):
            try:
                print('[실사상세 입력값]', master_id, idx, item.get('itemCode'), safe_str(item.get('itemName')), safe_str(item.get('spec')), safe_str(item.get('unit')), safe_float(item.get('systemQty')), safe_float(item.get('countedQty')), safe_float(item.get('diffQty')), safe_str(item.get('rowRemarks')))
                cursor.execute('''
                    INSERT INTO inventory_audit_detail (master_id, seq_no, item_code, item_name, spec, unit, system_qty, counted_qty, diff_qty, row_remarks)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    master_id,
                    idx,
                    item.get('itemCode'),
                    safe_str(item.get('itemName')),
                    safe_str(item.get('spec')),
                    safe_str(item.get('unit')),
                    safe_float(item.get('systemQty')),
                    safe_float(item.get('countedQty')),
                    safe_float(item.get('diffQty')),
                    safe_str(item.get('rowRemarks'))
                ))
            except Exception as e:
                print('[실사상세 INSERT 오류]', str(e), '입력값:', master_id, idx, item.get('itemCode'), safe_str(item.get('itemName')), safe_str(item.get('spec')), safe_str(item.get('unit')), safe_float(item.get('systemQty')), safe_float(item.get('countedQty')), safe_float(item.get('diffQty')), safe_str(item.get('rowRemarks')))
                conn.rollback()
                return jsonify({'success': False, 'message': '실사 상세 저장에 실패했습니다.'})
        conn.commit()
        print('[실사등록 최종 성공] master_id:', master_id)
        return jsonify({'success': True, 'message': '실사 등록이 완료되었습니다.', 'auditId': master_id})
    except Exception as e:
        print('[실사등록 전체 오류]', str(e))
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)})

@inventory.route('/api/inventory/audit/list')
@login_required
def inventory_audit_list_api():
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    warehouse_code = request.args.get('searchWarehouse')
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    offset = (page - 1) * page_size
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    # 전체 행 수 구하기
    count_query = '''
        SELECT COUNT(*)
        FROM inventory_audit_master m
        LEFT JOIN code_warehouse w ON m.warehouse_code = w.WH_CODE
        WHERE 1=1
    '''
    params = []
    if start_date:
        count_query += ' AND m.audit_date >= ?'
        params.append(start_date)
    if end_date:
        count_query += ' AND m.audit_date <= ?'
        params.append(end_date)
    if warehouse_code:
        count_query += ' AND m.warehouse_code = ?'
        params.append(warehouse_code)
    cursor.execute(count_query, params)
    total_count = cursor.fetchone()[0]
    # 데이터 쿼리
    query = '''
        SELECT m.id, m.audit_date, m.warehouse_code, w.WH_NAME as warehouse_name, m.remarks, m.createuser, m.createdate, m.status
        FROM inventory_audit_master m
        LEFT JOIN code_warehouse w ON m.warehouse_code = w.WH_CODE
        WHERE 1=1
    '''
    params2 = list(params)
    if start_date:
        query += ' AND m.audit_date >= ?'
    if end_date:
        query += ' AND m.audit_date <= ?'
    if warehouse_code:
        query += ' AND m.warehouse_code = ?'
    query += ' ORDER BY m.audit_date DESC, m.id DESC'
    query += ' OFFSET ? ROWS FETCH NEXT ? ROWS ONLY'
    params2.extend([offset, page_size])
    cursor.execute(query, params2)
    rows = cursor.fetchall()
    result = []
    for idx, r in enumerate(rows):
        # 전체 내림차순 순번 (최신이 1번)
        row_num = total_count - offset - idx
        result.append({
            'id': r.id,
            'audit_date': r.audit_date.strftime('%Y-%m-%d') if r.audit_date else '',
            'warehouse_code': r.warehouse_code,
            'warehouse_name': r.warehouse_name or '',
            'remarks': r.remarks or '',
            'createuser': r.createuser or '',
            'createdate': r.createdate.strftime('%Y-%m-%d %H:%M:%S') if r.createdate else '',
            'row_num': row_num,
            'status': r.status or 'N'
        })
    return jsonify({'totalCount': total_count, 'data': result})

@inventory.route('/api/inventory/audit/delete/<int:audit_id>', methods=['DELETE'])
@login_required
def delete_inventory_audit(audit_id):
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        
        # 먼저 상세 데이터 삭제
        cursor.execute('DELETE FROM inventory_audit_detail WHERE master_id = ?', (audit_id,))
        
        # 마스터 데이터 삭제
        cursor.execute('DELETE FROM inventory_audit_master WHERE id = ?', (audit_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '삭제되었습니다.'})
    except Exception as e:
        print(f"삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/detail/<int:audit_id>')
@login_required
def inventory_audit_detail(audit_id):
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        
        # 마스터 데이터 조회 (창고명 포함)
        cursor.execute('''
            SELECT m.id, m.audit_date, m.warehouse_code, m.remarks, m.createuser, m.createdate, m.status,
                   w.WH_NAME as warehouse_name
            FROM inventory_audit_master m
            LEFT JOIN code_warehouse w ON m.warehouse_code = w.WH_CODE
            WHERE m.id = ?
        ''', (audit_id,))
        master_row = cursor.fetchone()
        
        if not master_row:
            return jsonify({'success': False, 'message': '실사 데이터를 찾을 수 없습니다.'}), 404
        
        # 마스터 데이터 변환
        data = {
            'id': master_row.id,
            'audit_date': master_row.audit_date.strftime('%Y-%m-%d') if master_row.audit_date else '',
            'warehouse_code': master_row.warehouse_code,
            'warehouse_name': master_row.warehouse_name,
            'remarks': master_row.remarks,
            'createuser': master_row.createuser,
            'createdate': master_row.createdate.strftime('%Y-%m-%d %H:%M:%S') if master_row.createdate else '',
            'status': master_row.status
        }
        
        conn.close()
        return jsonify({'success': True, 'data': data})
        
    except Exception as e:
        print(f"상세 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/items/<int:audit_id>')
@login_required
def inventory_audit_items(audit_id):
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        
        # 상세 데이터 조회 (품목마스터 정보 포함)
        cursor.execute('''
            SELECT 
                d.seq_no, 
                d.item_code, 
                d.item_name, 
                d.spec, 
                d.unit, 
                d.system_qty, 
                d.counted_qty, 
                d.diff_qty, 
                d.row_remarks,
                m.ItemName as master_item_name,
                m.Spec as master_spec,
                m.unit as master_unit,
                m.purchasePrice,
                m.salesprice,
                m.usage,
                m.remarks as master_remarks
            FROM inventory_audit_detail d
            LEFT JOIN ItemMaster m ON d.item_code = m.ItemCode
            WHERE d.master_id = ?
            ORDER BY d.seq_no
        ''', (audit_id,))
        detail_rows = cursor.fetchall()
        
        # 상세 데이터 변환
        data = []
        for row in detail_rows:
            # 품목명은 실사상세에 있는 것을 우선, 없으면 마스터에서 가져오기
            item_name = row.item_name if row.item_name else row.master_item_name
            spec = row.spec if row.spec else row.master_spec
            unit = row.unit if row.unit else row.master_unit
            
            data.append({
                'seq_no': row.seq_no,
                'item_code': row.item_code,
                'item_name': item_name,
                'spec': spec,
                'unit': unit,
                'purchase_price': float(row.purchasePrice) if row.purchasePrice is not None else 0.0,
                'sales_price': float(row.salesprice) if row.salesprice is not None else 0.0,
                'usage': row.usage,
                'master_remarks': row.master_remarks,
                'system_qty': float(row.system_qty) if row.system_qty is not None else 0.0,
                'audit_qty': float(row.counted_qty) if row.counted_qty is not None else 0.0,
                'diff_qty': float(row.diff_qty) if row.diff_qty is not None else 0.0,
                'remarks': row.row_remarks
            })
        
        conn.close()
        return jsonify({'success': True, 'data': data})
        
    except Exception as e:
        print(f"상세 항목 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/update', methods=['PUT'])
@login_required
def inventory_audit_update():
    data = request.get_json()
    audit_id = data.get('auditId')
    audit_date = data.get('auditDate')
    warehouse_code = data.get('warehouseCode')
    remarks = safe_str(data.get('remarks'))
    items = data.get('items', [])
    username = current_user.username if hasattr(current_user, 'username') else 'system'
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    
    try:
        # 기존 상세 데이터 삭제
        cursor.execute('DELETE FROM inventory_audit_detail WHERE master_id = ?', (audit_id,))
        
        # 마스터 데이터 업데이트
        cursor.execute('''
            UPDATE inventory_audit_master 
            SET audit_date = ?, warehouse_code = ?, remarks = ?
            WHERE id = ?
        ''', (audit_date, warehouse_code, remarks, audit_id))
        
        # 상세 데이터 재저장
        for idx, item in enumerate(items, 1):
            cursor.execute('''
                INSERT INTO inventory_audit_detail (master_id, seq_no, item_code, item_name, spec, unit, system_qty, counted_qty, diff_qty, row_remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                audit_id,
                idx,
                item.get('itemCode'),
                safe_str(item.get('itemName')),
                safe_str(item.get('spec')),
                safe_str(item.get('unit')),
                safe_float(item.get('systemQty')),
                safe_float(item.get('countedQty')),
                safe_float(item.get('diffQty')),
                safe_str(item.get('rowRemarks'))
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '실사 수정이 완료되었습니다.'})
        
    except Exception as e:
        print(f"수정 오류: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/apply/<int:audit_id>', methods=['POST'])
@login_required
def apply_inventory_audit(audit_id):
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        # 1. 마스터/상세 데이터 조회
        cursor.execute('SELECT id, warehouse_code, audit_date FROM inventory_audit_master WHERE id = ?', (audit_id,))
        master = cursor.fetchone()
        if not master:
            return jsonify({'success': False, 'message': '실사 정보를 찾을 수 없습니다.'}), 404
        warehouse_code = master.warehouse_code
        audit_date = master.audit_date
        cursor.execute('''
            SELECT item_code, counted_qty, system_qty
            FROM inventory_audit_detail
            WHERE master_id = ?
        ''', (audit_id,))
        details = cursor.fetchall()
        # 2. 각 품목별로 inventory 테이블 반영
        for row in details:
            item_code = row.item_code
            counted_qty = float(row.counted_qty or 0)
            system_qty = float(row.system_qty or 0)
            # inventory에 존재 여부 확인
            cursor.execute('SELECT InventoryId FROM Inventory WHERE WarehouseCode = ? AND ItemCode = ?', (warehouse_code, item_code))
            inv_row = cursor.fetchone()
            if inv_row:
                # 있으면 update
                cursor.execute('UPDATE Inventory SET CurrentStock = ? WHERE InventoryId = ?', (counted_qty, inv_row.InventoryId))
            else:
                # 없으면 insert
                cursor.execute('''
                    INSERT INTO Inventory (WarehouseCode, ItemCode, CurrentStock) VALUES (?, ?, ?)
                ''', (warehouse_code, item_code, counted_qty))
            # 3. inventory_transaction에 수불이력 추가
            diff_qty = counted_qty - system_qty
            in_qty = diff_qty if diff_qty > 0 else 0
            out_qty = -diff_qty if diff_qty < 0 else 0
            cursor.execute('''
                INSERT INTO Inventory_Transaction (
                    TransDate, WarehouseCode, ItemCode, TransType, InQty, OutQty, BalanceQty, Remarks, CreateUser, CreateDate
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                audit_date,
                warehouse_code,
                item_code,
                'T99',
                in_qty,
                out_qty,
                counted_qty,
                '재고실사 반영',
                current_user.username if hasattr(current_user, 'username') else 'system',
                audit_date
            ))
        # 4. 마스터 status를 'Y'로 변경
        cursor.execute('UPDATE inventory_audit_master SET status = ? WHERE id = ?', ('Y', audit_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '재고실사 반영이 완료되었습니다.'})
    except Exception as e:
        print(f"재고실사 반영 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/cancel_apply/<int:audit_id>', methods=['POST'])
@login_required
def cancel_apply_inventory_audit(audit_id):
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        # 1. 마스터 정보 조회
        cursor.execute('SELECT warehouse_code, audit_date FROM inventory_audit_master WHERE id = ?', (audit_id,))
        master = cursor.fetchone()
        if not master:
            return jsonify({'success': False, 'message': '실사 정보를 찾을 수 없습니다.'}), 404
        warehouse_code = master.warehouse_code
        audit_date = master.audit_date
        # 2. 상세 품목 목록 조회
        cursor.execute('SELECT item_code FROM inventory_audit_detail WHERE master_id = ?', (audit_id,))
        items = cursor.fetchall()
        # 3. inventory_transaction에서 해당 실사(T99) 이력 삭제 및 inventory 현재고 복원
        for row in items:
            item_code = row.item_code
            # transaction 삭제
            cursor.execute('''
                DELETE FROM Inventory_Transaction
                WHERE TransType = 'T99' AND WarehouseCode = ? AND ItemCode = ? AND TransDate = ?
            ''', (warehouse_code, item_code, audit_date))
            # system_qty 조회
            cursor.execute('''
                SELECT system_qty FROM inventory_audit_detail WHERE master_id = ? AND item_code = ?
            ''', (audit_id, item_code))
            system_qty_row = cursor.fetchone()
            system_qty = float(system_qty_row.system_qty) if system_qty_row and system_qty_row.system_qty is not None else 0.0
            # inventory에 존재 여부 확인
            cursor.execute('SELECT InventoryId FROM Inventory WHERE WarehouseCode = ? AND ItemCode = ?', (warehouse_code, item_code))
            inv_row = cursor.fetchone()
            if inv_row:
                # 있으면 update
                cursor.execute('UPDATE Inventory SET CurrentStock = ? WHERE InventoryId = ?', (system_qty, inv_row.InventoryId))
            else:
                # 없으면 insert
                cursor.execute('INSERT INTO Inventory (WarehouseCode, ItemCode, CurrentStock) VALUES (?, ?, ?)', (warehouse_code, item_code, system_qty))
        # 4. 마스터 status 'N'으로 변경
        cursor.execute('UPDATE inventory_audit_master SET status = ? WHERE id = ?', ('N', audit_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '재고실사 반영이 취소되었습니다.'})
    except Exception as e:
        print(f'반영취소 오류: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/edit/<int:audit_id>')
@login_required
def inventory_audit_edit_api(audit_id):
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        # 마스터 데이터 조회
        cursor.execute('''
            SELECT id, audit_date, warehouse_code, remarks, status
            FROM inventory_audit_master
            WHERE id = ?
        ''', (audit_id,))
        master = cursor.fetchone()
        if not master:
            return jsonify({'success': False, 'message': '실사 데이터를 찾을 수 없습니다.'}), 404
        master_data = {
            'id': master.id,
            'audit_date': master.audit_date.strftime('%Y-%m-%d') if master.audit_date else '',
            'warehouse_code': master.warehouse_code,
            'remarks': master.remarks,
            'status': master.status
        }
        # 상세 데이터 조회
        cursor.execute('''
            SELECT seq_no, item_code, item_name, spec, unit, system_qty, counted_qty, diff_qty, row_remarks
            FROM inventory_audit_detail
            WHERE master_id = ?
            ORDER BY seq_no
        ''', (audit_id,))
        details = []
        for row in cursor.fetchall():
            details.append({
                'seq_no': row.seq_no,
                'item_code': row.item_code,
                'item_name': row.item_name,
                'spec': row.spec,
                'unit': row.unit,
                'system_qty': float(row.system_qty) if row.system_qty is not None else 0.0,
                'counted_qty': float(row.counted_qty) if row.counted_qty is not None else 0.0,
                'diff_qty': float(row.diff_qty) if row.diff_qty is not None else 0.0,
                'row_remarks': row.row_remarks
            })
        conn.close()
        return jsonify({'success': True, 'master': master_data, 'details': details})
    except Exception as e:
        print(f"수정용 상세 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/detail_list')
@login_required
def inventory_audit_detail_list_api():
    try:
        item_code = request.args.get('item_code', '').strip()
        item_name = request.args.get('item_name', '').strip()
        warehouse_code = request.args.get('warehouse_code', '').strip()
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        query = '''
            SELECT 
                m.id as audit_id,
                m.audit_date,
                m.warehouse_code,
                w.WH_NAME as warehouse_name,
                m.remarks,
                d.item_code,
                d.item_name,
                d.spec,
                d.unit,
                d.system_qty,
                d.counted_qty,
                d.diff_qty,
                m.createdate,
                m.createuser
            FROM inventory_audit_master m
            LEFT JOIN code_warehouse w ON m.warehouse_code = w.WH_CODE
            JOIN inventory_audit_detail d ON m.id = d.master_id
            WHERE 1=1
        '''
        params = []
        if item_code:
            query += ' AND d.item_code LIKE ?'
            params.append(f'%{item_code}%')
        if item_name:
            query += ' AND d.item_name LIKE ?'
            params.append(f'%{item_name}%')
        if warehouse_code:
            query += ' AND m.warehouse_code = ?'
            params.append(warehouse_code)
        query += ' ORDER BY m.audit_date DESC, m.id DESC, d.seq_no ASC'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        data = []
        for r in rows:
            data.append({
                'audit_id': r.audit_id,
                'audit_date': r.audit_date.strftime('%Y-%m-%d') if r.audit_date else '',
                'warehouse_code': r.warehouse_code,
                'warehouse_name': r.warehouse_name,
                'remarks': r.remarks,
                'item_code': r.item_code,
                'item_name': r.item_name,
                'spec': r.spec,
                'unit': r.unit,
                'system_qty': float(r.system_qty) if r.system_qty is not None else 0.0,
                'counted_qty': float(r.counted_qty) if r.counted_qty is not None else 0.0,
                'diff_qty': float(r.diff_qty) if r.diff_qty is not None else 0.0,
                'createdate': r.createdate.strftime('%Y-%m-%d %H:%M:%S') if r.createdate else '',
                'createuser': r.createuser
            })
        conn.close()
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"현황 상세 리스트 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/audit/detail_list_excel')
@login_required
def inventory_audit_detail_list_excel():
    import io
    from openpyxl import Workbook
    from flask import send_file
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    try:
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        cursor.execute('''
            SELECT 
                m.id as audit_id,
                m.audit_date,
                m.warehouse_code,
                w.WH_NAME as warehouse_name,
                m.remarks,
                d.item_code,
                d.item_name,
                d.spec,
                d.unit,
                d.system_qty,
                d.counted_qty,
                d.diff_qty,
                m.createdate,
                m.createuser
            FROM inventory_audit_master m
            LEFT JOIN code_warehouse w ON m.warehouse_code = w.WH_CODE
            JOIN inventory_audit_detail d ON m.id = d.master_id
            ORDER BY m.audit_date DESC, m.id DESC, d.seq_no ASC
        ''')
        rows = cursor.fetchall()
        data = []
        for r in rows:
            data.append([
                None,  # 순번은 나중에
                r.audit_id,
                r.audit_date.strftime('%Y-%m-%d') if r.audit_date else '',
                r.warehouse_name or r.warehouse_code,
                r.remarks,
                r.item_code,
                r.item_name,
                r.spec,
                r.unit,
                float(r.system_qty) if r.system_qty is not None else 0.0,
                float(r.counted_qty) if r.counted_qty is not None else 0.0,
                float(r.diff_qty) if r.diff_qty is not None else 0.0,
                r.createdate.strftime('%Y-%m-%d %H:%M:%S') if r.createdate else '',
                r.createuser
            ])
        # 현재 페이지 데이터만
        page_data = data[start_idx:end_idx]
        # 순번 채우기
        for idx, row in enumerate(page_data):
            row[0] = start_idx + idx + 1
        # 엑셀 생성
        wb = Workbook()
        ws = wb.active
        ws.title = '재고실사현황상세'
        headers = ['순번','실사ID','실사일자','창고','비고','품목코드','품목명','규격','단위','시스템재고','실사수량','차이수량','등록일','등록자']
        ws.append(headers)
        for row in page_data:
            ws.append(row)
        # 엑셀 파일 반환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='재고실사현황상세.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"엑셀다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/stock_list')
@login_required
def inventory_stock_list_api():
    item_code = request.args.get('item_code', '').strip()
    item_name = request.args.get('item_name', '').strip()
    warehouse_code = request.args.get('warehouse_code', '').strip()
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
    cursor = conn.cursor()
    query = '''
     		select
		a.itemcode as item_code,
		b.itemname as item_name,
		b.spec as spec,
		b.unit as unit,

		a.warehousecode as warehouse_code,
		c.wh_name as warehouse_name,
		COALESCE(a.currentstock , 0) as current_stock
		from inventory a,itemmaster b,code_warehouse c
		where a.itemcode=b.itemcode
		and a.WarehouseCode=c.WH_CODE
        
    '''
    params = []
    if item_code:
        query += ' AND a.ItemCode LIKE ?'
        params.append(f'%{item_code}%')
    if item_name:
        query += ' AND b.ItemName LIKE ?'
        params.append(f'%{item_name}%')
    if warehouse_code:
        query += ' AND c.WH_CODE = ?'
        params.append(warehouse_code)
    query += ' ORDER BY a.ItemCode, c.WH_CODE'
    cursor.execute(query, params)
    rows = cursor.fetchall()
    data = []
    for r in rows:
        data.append({
            'item_code': r.item_code,
            'item_name': r.item_name,
            'spec': r.spec,
            'unit': r.unit,
            'warehouse_code': r.warehouse_code,
            'warehouse_name': r.warehouse_name,
            'current_stock': float(r.current_stock) if r.current_stock is not None else 0.0
        })
    conn.close()
    return jsonify({'success': True, 'data': data})

@inventory.route('/api/inventory/stock_list_excel')
@login_required
def inventory_stock_list_excel():
    import io
    from openpyxl import Workbook
    from flask import send_file
    
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    
    try:
        item_code = request.args.get('item_code', '').strip()
        item_name = request.args.get('item_name', '').strip()
        warehouse_code = request.args.get('warehouse_code', '').strip()
        
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        query = '''
            SELECT
                i.ItemCode as item_code,
                i.ItemName as item_name,
                i.Spec as spec,
                i.unit as unit,
                w.WH_CODE as warehouse_code,
                w.WH_NAME as warehouse_name,
                COALESCE(inv.CurrentStock, 0) as current_stock
            FROM ItemMaster i
            CROSS JOIN code_warehouse w
            LEFT JOIN Inventory inv ON i.ItemCode = inv.ItemCode AND w.WH_CODE = inv.WarehouseCode
            WHERE 1=1
        '''
        params = []
        if item_code:
            query += ' AND i.ItemCode LIKE ?'
            params.append(f'%{item_code}%')
        if item_name:
            query += ' AND i.ItemName LIKE ?'
            params.append(f'%{item_name}%')
        if warehouse_code:
            query += ' AND w.WH_CODE = ?'
            params.append(warehouse_code)
        query += ' ORDER BY i.ItemCode, w.WH_CODE'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        data = []
        for r in rows:
            data.append([
                None,  # 순번은 나중에
                r.item_code,
                r.item_name,
                r.warehouse_name or r.warehouse_code,
                r.spec,
                r.unit,
                float(r.current_stock) if r.current_stock is not None else 0.0
            ])
        
        # 현재 페이지 데이터만
        page_data = data[start_idx:end_idx]
        # 순번 채우기
        for idx, row in enumerate(page_data):
            row[0] = start_idx + idx + 1
        
        # 엑셀 생성
        wb = Workbook()
        ws = wb.active
        ws.title = '재고현황'
        headers = ['순번', '품목코드', '품명', '창고', '규격', '단위', '현재고']
        ws.append(headers)
        for row in page_data:
            ws.append(row)
        
        # 엑셀 파일 반환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='재고현황.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"재고현황 엑셀다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/transaction_list')
@login_required
def inventory_transaction_list_api():
    try:
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        item_code = request.args.get('item_code', '').strip()
        item_name = request.args.get('item_name', '').strip()
        warehouse_code = request.args.get('warehouse_code', '').strip()
        
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        
        query = '''
            SELECT 
                t.TransDate,
                t.ItemCode,
                i.ItemName,
                t.LotNo,
                t.WarehouseCode,
                w.WH_NAME as WarehouseName,
                t.TransType,
                ct.TR_NAME as TrName,
                t.InQty,
                t.OutQty,
                t.BalanceQty,
                t.Remarks,
                t.CreateUser,
                t.CreateDate
            FROM Inventory_Transaction t
            LEFT JOIN ItemMaster i ON t.ItemCode = i.ItemCode
            LEFT JOIN Code_Warehouse w ON t.WarehouseCode = w.WH_CODE
            LEFT JOIN CODE_TRANSACTION ct ON t.TransType = ct.TR_CODE
            WHERE 1=1
        '''
        params = []
        if start_date:
            query += ' AND t.TransDate >= ?'
            params.append(start_date)
        if end_date:
            query += ' AND t.TransDate <= ?'
            params.append(end_date)
        if item_code:
            query += ' AND t.ItemCode LIKE ?'
            params.append(f'%{item_code}%')
        if item_name:
            query += ' AND i.ItemName LIKE ?'
            params.append(f'%{item_name}%')
        if warehouse_code:
            query += ' AND t.WarehouseCode = ?'
            params.append(warehouse_code)
        query += ' ORDER BY t.TransDate DESC, t.ItemCode'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        data = []
        for r in rows:
            qty = r.InQty if r.TransType == 'IN' else (-r.OutQty if r.TransType == 'OUT' else r.InQty - r.OutQty)
            data.append({
                'transaction_date': r.TransDate.strftime('%Y-%m-%d') if r.TransDate else '',
                'item_code': r.ItemCode,
                'item_name': r.ItemName,
                'lotno': r.LotNo,
                'warehouse_code': r.WarehouseCode,
                'warehouse_name': r.WarehouseName,
                'transaction_type': r.TransType,
                'tr_name': r.TrName,
                'in_qty': float(r.InQty) if r.InQty is not None else 0.0,
                'out_qty': float(r.OutQty) if r.OutQty is not None else 0.0,
                'balance_qty': float(r.BalanceQty) if r.BalanceQty is not None else 0.0,
                'quantity': qty,
                'unit': '',
                'remark': r.Remarks,
                'create_user': r.CreateUser,
                'create_date': r.CreateDate.strftime('%Y-%m-%d %H:%M:%S') if r.CreateDate else ''
            })
        conn.close()
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"수불현황 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inventory.route('/api/inventory/transaction_list_excel')
@login_required
def inventory_transaction_list_excel():
    import io
    from openpyxl import Workbook
    from flask import send_file
    
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    
    try:
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        item_code = request.args.get('item_code', '').strip()
        item_name = request.args.get('item_name', '').strip()
        warehouse_code = request.args.get('warehouse_code', '').strip()
        
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        
        query = '''
            SELECT 
                t.TransDate,
                t.ItemCode,
                i.ItemName,
                t.LotNo,
                t.WarehouseCode,
                w.WH_NAME as WarehouseName,
                t.TransType,
                ct.TR_NAME as TrName,
                t.InQty,
                t.OutQty,
                t.BalanceQty,
                t.Remarks,
                t.CreateUser,
                t.CreateDate
            FROM Inventory_Transaction t
            LEFT JOIN ItemMaster i ON t.ItemCode = i.ItemCode
            LEFT JOIN Code_Warehouse w ON t.WarehouseCode = w.WH_CODE
            LEFT JOIN CODE_TRANSACTION ct ON t.TransType = ct.TR_CODE
            WHERE 1=1
        '''
        params = []
        if start_date:
            query += ' AND t.TransDate >= ?'
            params.append(start_date)
        if end_date:
            query += ' AND t.TransDate <= ?'
            params.append(end_date)
        if item_code:
            query += ' AND t.ItemCode LIKE ?'
            params.append(f'%{item_code}%')
        if item_name:
            query += ' AND i.ItemName LIKE ?'
            params.append(f'%{item_name}%')
        if warehouse_code:
            query += ' AND t.WarehouseCode = ?'
            params.append(warehouse_code)
        query += ' ORDER BY t.TransDate DESC, t.ItemCode'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        data = []
        for r in rows:
            qty = r.InQty if r.TransType == 'IN' else (-r.OutQty if r.TransType == 'OUT' else r.InQty - r.OutQty)
            transaction_type_text = '입고' if r.TransType == 'IN' else '출고' if r.TransType == 'OUT' else '조정'
            data.append([
                None,  # 순번은 나중에
                r.TransDate.strftime('%Y-%m-%d') if r.TransDate else '',
                r.ItemCode,
                r.ItemName,
                r.LotNo,
                r.WarehouseName or r.WarehouseCode,
                r.TransType,
                r.TrName,
                transaction_type_text,
                float(r.InQty) if r.InQty is not None else 0.0,
                float(r.OutQty) if r.OutQty is not None else 0.0,
                qty,
                r.BalanceQty,
                '',
                r.Remarks or '',
                r.CreateUser,
                r.CreateDate.strftime('%Y-%m-%d %H:%M:%S') if r.CreateDate else ''
            ])
        # 현재 페이지 데이터만
        page_data = data[start_idx:end_idx]
        # 순번 채우기
        for idx, row in enumerate(page_data):
            row[0] = start_idx + idx + 1
        # 엑셀 생성
        wb = Workbook()
        ws = wb.active
        ws.title = '수불현황'
        headers = ['순번', '거래일자', '품목코드', '품명', 'LOT', '창고', '거래구분', '거래구분명', '거래구분(한글)', '입고수량', '출고수량', '수불수량', '잔여수량', '단위', '비고', '등록자', '등록일'];
        ws.append(headers)
        for row in page_data:
            ws.append(row)
        # 엑셀 파일 반환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='수불현황.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"수불현황 엑셀다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# register a root-level alias for this single API when the blueprint is registered
def _register_root_route(state):
    # state.app is the Flask application instance
    state.app.add_url_rule(
        '/api/warehouse/options',
        endpoint='warehouse_options_root',
        view_func=warehouse_options,
        methods=['GET']
    )

# ensure the alias is created when the blueprint is registered to the app
inventory.record(_register_root_route)

# register root-level aliases for selected inventory APIs (keep blueprint prefix for others)
def _register_api_aliases(state):
    app = state.app
    # avoid endpoint name collision by using unique endpoint names
    app.add_url_rule(
        '/api/inventory/item/options',
        endpoint='inventory_item_options_root',
        view_func=inventory_item_options,
        methods=['GET']
    )
    app.add_url_rule(
        '/api/inventory/item/detail/<item_code>',
        endpoint='inventory_item_detail_root',
        view_func=inventory_item_detail,
        methods=['GET']
    )

# ensure alias routes are created when blueprint is registered
inventory.record(_register_api_aliases)

# 입고관리(Inbound Manage)
@inventory.route('/inbound_manage')
def inbound_manage():
    # 입고관리 페이지 렌더링
    return render_template('inventory/inbound_manage.html')

# 출고관리(Outbound Manage)
@inventory.route('/outbound_manage')
def outbound_manage():
    # 출고관리 페이지 렌더링
    return render_template('inventory/outbound_manage.html')

@inventory.route('/api/inventory/inbound_transaction_list')
@login_required
def inbound_transaction_list():
    try:
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        item_code = request.args.get('item_code', '').strip()
        # note: item_name filter removed because ItemName/ItemDesc not returned
        warehouse_code = request.args.get('warehouse_code', '').strip()

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        query = """
            SELECT TransId,
                TransDate,
                ItemCode,
                LotNo,
                TransType,
                TransType2,
                RefNo,
                WarehouseCode,
                InQty,
                Remarks
            FROM Inventory_Transaction
            WHERE TransType = 'IN'
               AND 1=1
        """
        params = []
        if start_date:
            query += " AND TransDate >= ?"
            params.append(start_date)
        if end_date:
            query += " AND TransDate <= ?"
            params.append(end_date)
        if item_code:
            query += " AND ItemCode LIKE ?"
            params.append(f'%{item_code}%')
        if warehouse_code:
            query += " AND WarehouseCode = ?"
            params.append(warehouse_code)
        query += " ORDER BY TransDate DESC, TransId DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()
        data = []
        for r in rows:
            data.append({
                'TransId': r.TransId,
                'TransDate': r.TransDate.strftime('%Y-%m-%d') if r.TransDate else '',
                'ItemCode': r.ItemCode,
                'LotNo': r.LotNo,
                'TransType': r.TransType,
                'TransType2': getattr(r, 'TransType2', '') if getattr(r, 'TransType2', None) is not None else '',
                'RefNo': r.RefNo,
                'WarehouseCode': r.WarehouseCode,
                'InQty': float(r.InQty) if r.InQty is not None else 0.0,
                'Remarks': r.Remarks or ''
            })
        conn.close()
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"입고 이력 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500



@inventory.route('/api/inventory/inbound_transaction', methods=['POST'])
@login_required
def inbound_transaction_create():
    try:
        data = request.get_json() or {}
        # 필드 파싱 / 기본값
        trans_date = data.get('TransDate') or datetime.now().strftime('%Y-%m-%d')
        item_code = data.get('ItemCode', '').strip()
        warehouse_code = data.get('WarehouseCode', '').strip()
        lot_no = data.get('LotNo', '').strip()
        # DB에 저장되는 기본 거래구분은 'IN' 으로 고정
        trans_type = 'IN'
        # 실제 세부 코드(I01, I11, I31 등)는 TransType2에 저장
        trans_type2 = (data.get('TransType2') or data.get('TransType') or '').strip()
        ref_no = data.get('RefNo', '').strip()
        in_qty = safe_float(data.get('InQty', 0))
        out_qty = safe_float(data.get('OutQty', 0))
        remarks = safe_str(data.get('Remarks')) or ''

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()

        # 현재 잔액(마지막 트랜잭션 기준) 조회
        last_balance = 0.0
        try:
            cursor.execute('SELECT TOP 1 BalanceQty FROM Inventory_Transaction WHERE ItemCode = ? AND WarehouseCode = ? ORDER BY TransId DESC', (item_code, warehouse_code))
            row = cursor.fetchone()
            if row and getattr(row, 'BalanceQty', None) is not None:
                last_balance = float(row.BalanceQty)
        except Exception:
            last_balance = 0.0

        new_balance = last_balance + in_qty - out_qty
        create_user = current_user.username if hasattr(current_user, 'username') else 'system'
        create_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # INSERT (TransType2 컬럼 포함)
        cursor.execute('''
            INSERT INTO Inventory_Transaction
            (TransDate, WarehouseCode, ItemCode, LotNo, TransType, TransType2, RefNo, InQty, OutQty, BalanceQty, Remarks, CreateUser, CreateDate)
            OUTPUT INSERTED.TransId
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (trans_date, warehouse_code, item_code, lot_no, trans_type, trans_type2, ref_no, in_qty, out_qty, new_balance, remarks, create_user, create_date))

        row = cursor.fetchone()
        inserted_transid = row[0] if row and row[0] is not None else None
        conn.commit()

        inserted = {
            'TransId': inserted_transid,
            'TransDate': trans_date,
            'ItemCode': item_code,
            'LotNo': lot_no,
            'TransType': trans_type,
            'TransType2': trans_type2,
            'RefNo': ref_no,
            'WarehouseCode': warehouse_code,
            'InQty': in_qty
        }

        conn.close()
        return jsonify({'success': True, 'data': inserted})
    except Exception as e:
        print(f"입고 등록 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@inventory.route('/api/inventory/inbound_transaction_list_excel')
@login_required
def inbound_transaction_list_excel():
    import io
    from openpyxl import Workbook
    from flask import send_file

    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size

    try:
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        item_code = request.args.get('item_code', '').strip()
        warehouse_code = request.args.get('warehouse_code', '').strip()

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        query = """
            SELECT
                TransId,
                TransDate,
                ItemCode,
                LotNo,
                TransType,
                TransType2,
                RefNo,
                WarehouseCode,
                InQty,
                Remarks
            FROM Inventory_Transaction
            WHERE TransType = 'IN'
               AND 1=1
        """
        params = []
        if start_date:
            query += " AND TransDate >= ?"
            params.append(start_date)
        if end_date:
            query += " AND TransDate <= ?"
            params.append(end_date)
        if item_code:
            query += " AND ItemCode LIKE ?"
            params.append(f'%{item_code}%')
        if warehouse_code:
            query += " AND WarehouseCode = ?"
            params.append(warehouse_code)
        query += " ORDER BY TransDate DESC, TransId DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        data = []
        for r in rows:
            data.append([
                None,  # 순번 채우기
                r.TransId,
                r.TransDate.strftime('%Y-%m-%d') if r.TransDate else '',
                r.ItemCode,
                r.LotNo or '',
                r.TransType or '',
                getattr(r, 'TransType2', '') if getattr(r, 'TransType2', None) is not None else '',
                r.RefNo or '',
                r.WarehouseCode or '',
                float(r.InQty) if r.InQty is not None else 0.0,
                r.Remarks or ''
            ])

        # 페이지 적용
        page_data = data[start_idx:end_idx]
        for idx, row in enumerate(page_data):
            row[0] = start_idx + idx + 1

        wb = Workbook()
        ws = wb.active
        ws.title = '입고이력'
        headers = ['순번','TransId','거래일자','품목코드','LOT','거래구분','거래구분2','RefNo','창고','입고수량','비고']
        ws.append(headers)
        for row in page_data:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='입고이력.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"입고 이력 엑셀다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


    # --- Outbound transaction APIs (copied from inbound, adjusted for OUT) ---
@inventory.route('/api/inventory/outbound_transaction_list')
@login_required
def outbound_transaction_list():
    try:
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        item_code = request.args.get('item_code', '').strip()
        # note: item_name filter removed if not returned by query
        warehouse_code = request.args.get('warehouse_code', '').strip()

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        query = """
            SELECT TransId,
                TransDate,
                ItemCode,
                LotNo,
                TransType,
                TransType2,
                RefNo,
                WarehouseCode,
                OutQty,
                Remarks
            FROM Inventory_Transaction
            WHERE TransType = 'OUT'
               AND 1=1
        """
        params = []
        if start_date:
            query += " AND TransDate >= ?"
            params.append(start_date)
        if end_date:
            query += " AND TransDate <= ?"
            params.append(end_date)
        if item_code:
            query += " AND ItemCode LIKE ?"
            params.append(f'%{item_code}%')
        if warehouse_code:
            query += " AND WarehouseCode = ?"
            params.append(warehouse_code)
        query += " ORDER BY TransDate DESC, TransId DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()
        data = []
        for r in rows:
            data.append({
                'TransId': r.TransId,
                'TransDate': r.TransDate.strftime('%Y-%m-%d') if r.TransDate else '',
                'ItemCode': r.ItemCode,
                'LotNo': r.LotNo,
                'TransType': r.TransType,
                'TransType2': getattr(r, 'TransType2', '') if getattr(r, 'TransType2', None) is not None else '',
                'RefNo': r.RefNo,
                'WarehouseCode': r.WarehouseCode,
                'OutQty': float(r.OutQty) if r.OutQty is not None else 0.0,
                'Remarks': r.Remarks or ''
            })
        conn.close()
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"출고 이력 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@inventory.route('/api/inventory/outbound_transaction', methods=['POST'])
@login_required
def outbound_transaction_create():
    try:
        data = request.get_json() or {}
        # 필드 파싱 / 기본값
        trans_date = data.get('TransDate') or datetime.now().strftime('%Y-%m-%d')
        item_code = data.get('ItemCode', '').strip()
        warehouse_code = data.get('WarehouseCode', '').strip()
        lot_no = data.get('LotNo', '').strip()
        # DB에 저장되는 기본 거래구분은 'OUT' 으로 고정
        trans_type = 'OUT'
        # 실제 세부 코드(I01, O11 등)는 TransType2에 저장
        trans_type2 = (data.get('TransType2') or data.get('TransType') or '').strip()
        ref_no = data.get('RefNo', '').strip()
        in_qty = safe_float(data.get('InQty', 0))
        out_qty = safe_float(data.get('OutQty', 0))
        remarks = safe_str(data.get('Remarks')) or ''

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()

        # 현재 잔액(마지막 트랜잭션 기준) 조회
        last_balance = 0.0
        try:
            cursor.execute('SELECT TOP 1 BalanceQty FROM Inventory_Transaction WHERE ItemCode = ? AND WarehouseCode = ? ORDER BY TransId DESC', (item_code, warehouse_code))
            row = cursor.fetchone()
            if row and getattr(row, 'BalanceQty', None) is not None:
                last_balance = float(row.BalanceQty)
        except Exception:
            last_balance = 0.0

        new_balance = last_balance + in_qty - out_qty
        create_user = current_user.username if hasattr(current_user, 'username') else 'system'
        create_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # INSERT (TransType2 컬럼 포함)
        cursor.execute('''
            INSERT INTO Inventory_Transaction
            (TransDate, WarehouseCode, ItemCode, LotNo, TransType, TransType2, RefNo, InQty, OutQty, BalanceQty, Remarks, CreateUser, CreateDate)
            OUTPUT INSERTED.TransId
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (trans_date, warehouse_code, item_code, lot_no, trans_type, trans_type2, ref_no, in_qty, out_qty, new_balance, remarks, create_user, create_date))

        row = cursor.fetchone()
        inserted_transid = row[0] if row and row[0] is not None else None
        conn.commit()

        inserted = {
            'TransId': inserted_transid,
            'TransDate': trans_date,
            'ItemCode': item_code,
            'LotNo': lot_no,
            'TransType': trans_type,
            'TransType2': trans_type2,
            'RefNo': ref_no,
            'WarehouseCode': warehouse_code,
            'OutQty': out_qty
        }

        conn.close()
        return jsonify({'success': True, 'data': inserted})
    except Exception as e:
        print(f"출고 등록 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@inventory.route('/api/inventory/outbound_transaction_list_excel')
@login_required
def outbound_transaction_list_excel():
    import io
    from openpyxl import Workbook
    from flask import send_file

    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size

    try:
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        item_code = request.args.get('item_code', '').strip()
        warehouse_code = request.args.get('warehouse_code', '').strip()

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=118.67.132.208;DATABASE=BIGBOY;UID=brother;PWD=jobgate@m1n;')
        cursor = conn.cursor()
        query = """
            SELECT
                TransId,
                TransDate,
                ItemCode,
                LotNo,
                TransType,
                TransType2,
                RefNo,
                WarehouseCode,
                OutQty,
                Remarks
            FROM Inventory_Transaction
            WHERE TransType = 'OUT'
               AND 1=1
        """
        params = []
        if start_date:
            query += " AND TransDate >= ?"
            params.append(start_date)
        if end_date:
            query += " AND TransDate <= ?"
            params.append(end_date)
        if item_code:
            query += " AND ItemCode LIKE ?"
            params.append(f'%{item_code}%')
        if warehouse_code:
            query += " AND WarehouseCode = ?"
            params.append(warehouse_code)
        query += " ORDER BY TransDate DESC, TransId DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        data = []
        for r in rows:
            data.append([
                None,  # 순번 채우기
                r.TransId,
                r.TransDate.strftime('%Y-%m-%d') if r.TransDate else '',
                r.ItemCode,
                r.LotNo or '',
                r.TransType or '',
                getattr(r, 'TransType2', '') if getattr(r, 'TransType2', None) is not None else '',
                r.RefNo or '',
                r.WarehouseCode or '',
                float(r.OutQty) if r.OutQty is not None else 0.0,
                r.Remarks or ''
            ])

        # 페이지 적용
        page_data = data[start_idx:end_idx]
        for idx, row in enumerate(page_data):
            row[0] = start_idx + idx + 1

        wb = Workbook()
        ws = wb.active
        ws.title = '출고이력'
        headers = ['순번','TransId','거래일자','품목코드','LOT','거래구분','거래구분2','RefNo','창고','출고수량','비고']
        ws.append(headers)
        for row in page_data:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='출고이력.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"출고 이력 엑셀다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500